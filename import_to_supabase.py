#!/usr/bin/env python3
"""
Import products from an Excel file into the Supabase `products` table.

Excel file: murph_products_v2.xlsx
Columns (in order):
    Name | Category | Rating | Price | Image URL | HipoBuy Link

Usage:
    pip install openpyxl requests
    python import_to_supabase.py

Reads Supabase credentials from environment variables, falling back to a
local .env.local file:
    NEXT_PUBLIC_SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY
"""

import os
import sys

import requests
from openpyxl import load_workbook

EXCEL_FILE = "murph_products_v2.xlsx"
TABLE = "products"
BATCH_SIZE = 100


def load_env_local(path: str = ".env.local") -> None:
    """Populate os.environ from a .env.local file if present."""
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


def get_credentials() -> tuple[str, str]:
    load_env_local()
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        sys.exit(
            "Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY. "
            "Set them as env vars or in .env.local."
        )
    return url, key


def parse_rating(value) -> float:
    try:
        return max(0.0, min(5.0, float(value)))
    except (TypeError, ValueError):
        return 4.5


def read_rows(path: str) -> list[dict]:
    if not os.path.exists(path):
        sys.exit(f"Excel file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    for idx, raw in enumerate(ws.iter_rows(values_only=True)):
        # Skip the header row.
        if idx == 0:
            continue
        if raw is None or all(cell is None for cell in raw):
            continue

        # Pad/truncate to the 6 expected columns.
        cells = list(raw)[:6] + [None] * (6 - len(raw))
        name, category, rating, price, image_url, link = cells

        name = (str(name).strip() if name is not None else "")
        link = (str(link).strip() if link is not None else "")
        if not name or not link:
            # Name and link are required by the table / API.
            continue

        rows.append(
            {
                "name": name,
                "category": (str(category).strip() if category else "Shoes"),
                "rating": parse_rating(rating),
                # Price is stored as-is (string), e.g. "¥299".
                "price": (str(price).strip() if price is not None else ""),
                "image": (str(image_url).strip() if image_url else ""),
                "link": link,
            }
        )

    wb.close()
    return rows


def insert_rows(url: str, key: str, rows: list[dict]) -> None:
    endpoint = f"{url}/rest/v1/{TABLE}"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    inserted = 0
    for start in range(0, len(rows), BATCH_SIZE):
        batch = rows[start : start + BATCH_SIZE]
        resp = requests.post(endpoint, headers=headers, json=batch, timeout=60)
        if not resp.ok:
            sys.exit(f"Insert failed ({resp.status_code}): {resp.text}")
        inserted += len(batch)
        print(f"Inserted {inserted}/{len(rows)}")


def main() -> None:
    url, key = get_credentials()
    rows = read_rows(EXCEL_FILE)
    if not rows:
        print("No valid rows found to import.")
        return
    print(f"Read {len(rows)} products from {EXCEL_FILE}. Importing...")
    insert_rows(url, key, rows)
    print("Done.")


if __name__ == "__main__":
    main()
